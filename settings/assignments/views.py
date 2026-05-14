from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
import openpyxl

from courses.models import Course
from accounts.models import User
from .models import (
    Assignment, Question, Submission, Answer,
    Topic, Trainer, Option
)

# ==========================================================
# STUDENT VIEWS
# ==========================================================

@login_required
def student_assignments(request):
    if request.user.role != "student":
        return redirect("login")

    assignments = Assignment.objects.filter(status="approved").order_by("-id")
    
    student_history = Submission.objects.filter(
        student=request.user
    ).select_related('assignment').order_by("-submitted_at")

    submission_map = {s.assignment_id: s for s in student_history}

    return render(request, "assignments/student_assignments.html", {
        "assignments": assignments,
        "student_history": student_history,
        "submission_map": submission_map
    })
@login_required
def start_exam(request, assignment_id):
    if request.user.role != "student":
        return redirect("login")

    assignment = get_object_or_404(Assignment, id=assignment_id, status="approved")
    
    submission, created = Submission.objects.get_or_create(
        assignment=assignment,
        student=request.user,
        defaults={"status": "started"}
    )

    if assignment.test_type == "mcq":
        questions = assignment.questions.all().prefetch_related('options') 
        return render(request, "assignments/start_exam.html", {
            "assignment": assignment,
            "questions": questions,
            "submission": submission
        })

    if assignment.test_type == "pdf":
        if not assignment.pdf_file:
            messages.error(request, "The PDF question paper is missing.")
            return redirect("assignments:student_assignments")
        return render(request, "assignments/pdf_exam.html", {
            "assignment": assignment,
            "submission": submission
        })

    if assignment.test_type == "gform":
        if not assignment.google_form_link:
            messages.error(request, "The Google Form link is missing.")
            return redirect("assignments:student_assignments")
        
        # ✅ Show an intermediate page instead of direct redirect
        return render(request, "assignments/gform_exam.html", {
            "assignment": assignment,
            "submission": submission
        })

    return redirect("assignments:student_assignments")
@login_required
def gform_exam(request, assignment_id):
    if request.user.role != "student":
        return redirect("login")
    assignment = get_object_or_404(Assignment, id=assignment_id, status="approved")
    submission, created = Submission.objects.get_or_create(
        assignment=assignment,
        student=request.user,
        defaults={"status": "started"}
    )
    return render(request, "assignments/gform_exam.html", {
        "assignment": assignment,
        "submission": submission
    })
@login_required
def submit_exam(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id, student=request.user)
    
    if request.method == "POST":
        if submission.assignment.test_type == "mcq":
            total_obtained = 0
            submission.answers.all().delete()
            
            questions = submission.assignment.questions.all()
            for q in questions:
                selected_id = request.POST.get(f"q_{q.id}")
                if selected_id:
                    option = Option.objects.filter(id=selected_id, question=q).first()
                    if option:
                        Answer.objects.create(
                            submission=submission,
                            question=q,
                            selected_option=option,
                            is_correct=option.is_correct
                        )
                        if option.is_correct:
                            total_obtained += q.marks
            
            submission.score = total_obtained
        
        submission.status = "submitted"
        submission.submitted_at = timezone.now()
        submission.save()
        
        messages.success(request, "Exam submitted successfully!")
        return redirect("assignments:exam_result", submission_id=submission.id)
    
    return redirect("assignments:student_assignments")

@login_required
def submit_pdf_exam(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    submission, _ = Submission.objects.get_or_create(assignment=assignment, student=request.user)
    if request.method == "POST":
        file = request.FILES.get("file_submission")
        if file:
            submission.pdf_answer = file
            submission.status = "submitted"
            submission.submitted_at = timezone.now()
            submission.save()
            return redirect("assignments:exam_result", submission_id=submission.id)
    return redirect("assignments:student_assignments")

@login_required
def submit_gform_exam(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    submission, _ = Submission.objects.get_or_create(assignment=assignment, student=request.user)
    if request.method == "POST":
        screenshot = request.FILES.get("screenshot")
        if screenshot:
            submission.pdf_answer = screenshot
            submission.status = "submitted"
            submission.submitted_at = timezone.now()
            submission.save()
            return redirect("assignments:exam_result", submission_id=submission.id)
    return redirect("assignments:student_assignments")
@login_required
def exam_result(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id)
    
    total_marks = submission.assignment.total_marks if submission.assignment.total_marks is not None else 100
    pass_marks = submission.assignment.pass_marks if submission.assignment.pass_marks is not None else 40
    
    percentage = 0
    passed = False
    pending_review = False  # NEW FLAG

    test_type = submission.assignment.test_type

    if test_type == "mcq":
        # MCQ: auto-graded, show pass/fail normally
        if submission.score is not None:
            if total_marks > 0:
                percentage = (submission.score / total_marks) * 100
            passed = submission.score >= pass_marks

    elif test_type in ["pdf", "gform"]:
        # PDF / Google Form: manually graded by trainer
        if submission.status == "graded" and submission.score is not None:
            # Trainer has graded it — now show real result
            if total_marks > 0:
                percentage = (submission.score / total_marks) * 100
            passed = submission.score >= pass_marks
        else:
            # Not yet graded — show pending
            pending_review = True

    return render(request, "assignments/exam_result.html", {
        "submission": submission,
        "assignment": submission.assignment,
        "total": total_marks,
        "pass_marks": pass_marks,
        "percentage": round(percentage, 2),
        "passed": passed,
        "pending_review": pending_review,  # NEW
        "test_type": test_type,            # NEW
    })

# ==========================================================
# TRAINER VIEWS
# ==========================================================

@login_required
def trainer_assignments(request):
    if request.user.role != "trainer":
        return redirect("login")
    
    trainer = getattr(request.user, "trainer_profile", None)
    
    trainer_assignments = Assignment.objects.filter(trainer=trainer).order_by("-id")
    trainer_submissions = Submission.objects.filter(assignment__trainer=trainer).select_related('student', 'assignment')
    
    student_ids = trainer_submissions.values_list('student', flat=True).distinct()
    assigned_students = User.objects.filter(id__in=student_ids)

    student_perf = []
    for student in assigned_students:
        s_stats = trainer_submissions.filter(student=student, status__in=['submitted', 'graded'])
        count = s_stats.count()
        avg_percent = sum(sub.percentage for sub in s_stats) / count if count > 0 else 0
        student_perf.append({
            "user": student,
            "exams_count": count,
            "percentage": round(avg_percent, 1),
        })

    return render(request, "assignments/trainer_assignments.html", {
        "assignments": trainer_assignments,
        "submissions": trainer_submissions.order_by('-submitted_at'),
        "student_perf": student_perf,
        "topics": Topic.objects.all(), 
        "courses": Course.objects.all(),
    })

@login_required
def trainer_create_assignment(request):
    trainer = getattr(request.user, "trainer_profile", None)
    if request.method == "POST":
        test_type = request.POST.get("test_type")
        total_marks = int(request.POST.get("total_marks") or 100)
        pass_marks = int(request.POST.get("pass_marks") or 40)

        try:
            with transaction.atomic():
                assignment = Assignment.objects.create(
                    title=request.POST.get("title"),
                    course_id=request.POST.get("course"),
                    topic_id=request.POST.get("topic"),
                    trainer=trainer,
                    created_by=request.user,
                    test_type=test_type,
                    total_marks=total_marks,
                    pass_marks=pass_marks,
                    status="approved"
                )

                if test_type == "mcq":
                    q_indices = [k.split('_')[1] for k in request.POST.keys() if k.startswith('q_') and k.endswith('_text')]
                    num_q = len(q_indices)
                    if num_q > 0:
                        marks_per_q = total_marks // num_q
                        for i in q_indices:
                            q_text = request.POST.get(f"q_{i}_text")
                            question = Question.objects.create(assignment=assignment, text=q_text, marks=marks_per_q)
                            j = 1
                            while request.POST.get(f"q_{i}_opt_{j}"):
                                Option.objects.create(
                                    question=question,
                                    text=request.POST.get(f"q_{i}_opt_{j}"),
                                    is_correct=(request.POST.get(f"q_{i}_correct") == str(j))
                                )
                                j += 1
                    # ✅ Removed update_total_marks() to preserve frontend value

                elif test_type == "pdf":
                    assignment.pdf_file = request.FILES.get("pdf_file")
                    assignment.save()

                elif test_type == "gform":
                    assignment.google_form_link = request.POST.get("google_form_link")
                    assignment.save()

            messages.success(request, f"Assignment '{assignment.title}' published!")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            
    return redirect("assignments:trainer_assignments")

@login_required
def trainer_grade_submission(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id)
    
    if submission.assignment.trainer.user != request.user and request.user.role != 'admin':
        messages.error(request, "Unauthorized access.")
        return redirect("assignments:trainer_assignments")

    if request.method == "POST":
        try:
            score = int(request.POST.get("score") or 0)
            submission.score = score
            submission.status = "graded"
            submission.save()
            messages.success(request, f"Successfully graded {submission.student.username}")
        except ValueError:
            messages.error(request, "Invalid score entered.")
        return redirect("assignments:trainer_assignments")
    
    return render(request, "assignments/grade_submission.html", {"submission": submission})

@login_required
def export_trainer_report(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Student", "Score", "Status", "Submitted At"])
    for s in Submission.objects.filter(assignment=assignment).select_related("student"):
        ws.append([s.student.username, s.score, s.status, s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else "N/A"])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="report_{assignment_id}.xlsx"'
    wb.save(response)
    return response

@login_required
def trainer_progress(request):
    trainer_assignments = Assignment.objects.filter(trainer__user=request.user)
    submissions = Submission.objects.filter(
        assignment__in=trainer_assignments
    ).select_related('student', 'assignment')

    student_perf = {}
    for s in submissions:
        sid = s.student.id
        if sid not in student_perf:
            student_perf[sid] = {
                'user': s.student,
                'total_score': 0,
                'max_marks': 0,
                'exams_count': 0
            }
        if s.score is not None:
            student_perf[sid]['total_score'] += s.score
            student_perf[sid]['max_marks'] += s.assignment.total_marks if s.assignment.total_marks is not None else 100
            student_perf[sid]['exams_count'] += 1

    for data in student_perf.values():
        if data['max_marks'] > 0:
            data['percentage'] = round((data['total_score'] / data['max_marks']) * 100, 2)
        else:
            data['percentage'] = 0

    return render(request, "dashboard/trainer_progress.html", {
        "student_perf": student_perf.values()
    })

# ==========================================================
# ADMIN VIEWS
# ==========================================================

@login_required
def admin_assignments(request):
    if request.user.role != "admin":
        return redirect("login")

    all_assignments = Assignment.objects.all().order_by("-id")
    all_submissions = Submission.objects.all().select_related('student', 'assignment', 'assignment__trainer')
    
    all_students = User.objects.filter(role="student")
    student_perf = []

    for student in all_students:
        s_stats = all_submissions.filter(student=student, status__in=['submitted', 'graded'])
        count = s_stats.count()
        avg_pct = sum(sub.percentage for sub in s_stats) / count if count > 0 else 0
        student_perf.append({
            "user": student,
            "exams_count": count,
            "percentage": round(avg_pct, 1),
        })

    return render(request, "assignments/admin_assignments.html", {
        "assignments": all_assignments,
        "student_perf": student_perf,
        "submissions": all_submissions.order_by("-submitted_at"),
        "courses": Course.objects.all(), 
        "trainers": Trainer.objects.all(), 
        "topics": Topic.objects.all()
    })

@login_required
def admin_create_assignment(request):
    if getattr(request.user, 'role', None) != "admin":
        return redirect("login")

    if request.method == "POST":
        try:
            with transaction.atomic():
                test_type = request.POST.get("test_type")
                total_marks = int(request.POST.get("total_marks") or 100)
                pass_marks = int(request.POST.get("pass_marks") or 40)
                duration = int(request.POST.get("duration") or 30)

                assignment = Assignment.objects.create(
                    title=request.POST.get("title", "").strip(),
                    description=request.POST.get("description", ""),
                    course_id=request.POST.get("course"),
                    trainer_id=request.POST.get("trainer"),
                    topic_id=request.POST.get("topic") or None,
                    created_by=request.user,
                    test_type=test_type,
                    status="approved",
                    duration=duration,
                    total_marks=total_marks,
                    pass_marks=pass_marks
                )

                if test_type == "mcq":
                    raw_q_indices = []
                    i = 1
                    while request.POST.get(f"q_{i}_text"):
                        raw_q_indices.append(i)
                        i += 1
                    
                    num_questions = len(raw_q_indices)
                    marks_per_q = total_marks // num_questions if num_questions > 0 else 0

                    for idx in raw_q_indices:
                        q_text = request.POST.get(f"q_{idx}_text")
                        time = request.POST.get(f"q_{idx}_time", 30)    #time in seconds
                        question = Question.objects.create(
                            assignment=assignment,
                            text=q_text,
                            marks=marks_per_q,
                            time_limit=time         #time in seconds
                        )
                        correct_val = request.POST.get(f"q_{idx}_correct")
                        j = 1
                        while request.POST.get(f"q_{idx}_opt_{j}"):
                            Option.objects.create(
                                question=question,
                                text=request.POST.get(f"q_{idx}_opt_{j}"),
                                is_correct=(str(j) == str(correct_val))
                            )
                            j += 1

                elif test_type == "pdf":
                    assignment.pdf_file = request.FILES.get("pdf_file")
                    assignment.save()
                
                elif test_type == "gform":
                    assignment.google_form_link = request.POST.get("google_form_link")
                    assignment.save()

            messages.success(request, f"✅ Assignment '{assignment.title}' published successfully!")
        except Exception as e:
            messages.error(request, f"❌ Error creating assignment: {str(e)}")
    
    return redirect("assignments:admin_assignments")

@login_required
def admin_view_results(request, assignment_id):
    if request.user.role != "admin":
        return redirect("login")
    assignment = get_object_or_404(Assignment, id=assignment_id)
    submissions = Submission.objects.filter(assignment=assignment).select_related('student')
    return render(request, "assignments/admin_view_results.html", {
        "assignment": assignment,
        "submissions": submissions
    })

@login_required
def admin_export(request):
    if request.user.role != "admin":
        return redirect("login")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Assignment", "Student", "Trainer", "Score", "Status", "Submitted At"])
    submissions = Submission.objects.all().select_related("assignment", "student", "assignment__trainer")
    for s in submissions:
        ws.append([
            s.assignment.title, s.student.username,
            s.assignment.trainer.user.username if s.assignment.trainer else "N/A",
            s.score, s.status,
            s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else "N/A"
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="admin_report.xlsx"'
    wb.save(response)
    return response

@login_required
def approve_assignment(request, pk):
    if request.user.role != "admin":
        return redirect("login")
    Assignment.objects.filter(pk=pk).update(status="approved")
    return redirect("assignments:admin_assignments")

@login_required
def reject_assignment(request, pk):
    if request.user.role != "admin":
        return redirect("login")
    Assignment.objects.filter(pk=pk).update(status="rejected")
    return redirect("assignments:admin_assignments")

# ==========================================================
# UTILITIES
# ==========================================================

@login_required
def add_question(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    if request.method == "POST":
        q = Question.objects.create(
            assignment=assignment,
            text=request.POST.get("text"),
            marks=request.POST.get("marks") or 1
        )
        for i in range(1, 5):
            opt = request.POST.get(f"option_{i}")
            if opt:
                Option.objects.create(question=q, text=opt, is_correct=(request.POST.get("correct") == str(i)))
        return redirect("assignments:trainer_assignments")
    return render(request, "assignments/add_question.html", {"assignment": assignment})

@login_required
def trainer_submissions(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    return render(request, "assignments/trainer_submissions.html", {
        "assignment": assignment,
        "submissions": Submission.objects.filter(assignment=assignment)
    })

@login_required
def leaderboard(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    submissions = Submission.objects.filter(assignment=assignment, status="submitted").order_by("-score", "submitted_at")
    return render(request, "assignments/leaderboard.html", {"assignment": assignment, "submissions": submissions})

@login_required
def certificate(request, submission_id):
    return HttpResponse(f"Certificate for submission {submission_id}")

@login_required
def export_all_assignments(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Title', 'Course', 'Topic', 'Test Type', 'Status'])
    for a in Assignment.objects.all():
        ws.append([a.title, str(a.course), str(a.topic), a.test_type, a.status])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="assignments.xlsx"'
    wb.save(response)
    return response

@login_required
def export_all_progress(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Student', 'Assignment', 'Score', 'Status'])
    for s in Submission.objects.all():
        ws.append([s.student.username, s.assignment.title, s.score, s.status])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="progress.xlsx"'
    wb.save(response)
    return response